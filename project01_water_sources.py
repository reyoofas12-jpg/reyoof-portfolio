import pandas as pd
import re
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

INPUT_FILE  = 'مصادر المياة بايثون.xlsx'
OUTPUT_FILE = 'النتيجة.xlsx'

LEAVE_STATUSES = {
    'إجازة مرضية','اجازة مرضية','إجازة رسمية','إجازة سنوية',
    'إجازة اختبارات','إجازة اجازة حج','استئذان كامل',
    'إجازة وفاة زوج، اصول او فروع'
}
REST_STATUSES    = {'يوم راحة'}
ABSENT_STATUSES  = {'غائب'}
PRESENT_STATUSES = {'حاضر','غير مكتمل'}

def parse_time(t):
    if not t or str(t).strip() in ('-','nan',''): return None
    t=str(t).strip().replace('صباحًا','AM').replace('مساءً','PM').replace('صباحا','AM').replace('مساءا','PM')
    m=re.match(r'(\d+):(\d+)\s*(AM|PM)',t,re.IGNORECASE)
    if not m: return None
    h,mn,p=int(m.group(1)),int(m.group(2)),m.group(3).upper()
    if p=='PM' and h!=12: h+=12
    elif p=='AM' and h==12: h=0
    return h*60+mn

def parse_shift_hours(s):
    if not s or str(s).strip()=='-': return None
    s=str(s).strip().replace('صباحًا','AM').replace('مساءً','PM').replace('صباحا','AM').replace('مساءا','PM')
    parts=re.split(r'\s*-\s*',s)
    if len(parts)!=2: return None
    st,en=parse_time(parts[0]),parse_time(parts[1])
    if st is None or en is None: return None
    d=en-st
    if d<0: d+=1440
    return round(d/60,2)

def parse_duration_mins(d):
    if not d or str(d).strip() in ('-','','nan','0:00'): return 0
    m=re.match(r'(\d+):(\d+)',str(d).strip())
    return int(m.group(1))*60+int(m.group(2)) if m else 0

def actual_hrs(entry,exit_):
    e,x=parse_time(entry),parse_time(exit_)
    if e is None or x is None: return None
    d=x-e
    if d<0: d+=1440
    return round(d/60,2)

def to_arabic_time(t):
    if not t: return '-'
    return t.replace('AM','صباحًا').replace('PM','مساءً')

def clean_required_hours(v, shift_str='-'):
    """تنظيف الساعات المطلوبة - إذا كانت 1 أو 2 احسبها من النوبة"""
    v=str(v).strip()
    if v in ('-','nan'): 
        sh = parse_shift_hours(shift_str)
        return str(int(sh)) if sh else '-'
    try:
        num = int(float(v))
        # إذا كانت 1 أو 2 استبدلها بساعات النوبة
        if num <= 2:
            sh = parse_shift_hours(shift_str)
            return str(int(sh)) if sh else str(num)
        return str(num)
    except:
        m=re.search(r'\d+',v)
        if m: 
            num = int(m.group())
            if num <= 2:
                sh = parse_shift_hours(shift_str)
                return str(int(sh)) if sh else str(num)
            return str(num)
        return '-'

def classify(status):
    if status in PRESENT_STATUSES: return 'present'
    if status in REST_STATUSES:    return 'rest'
    if status in LEAVE_STATUSES:   return 'leave'
    return 'absent'

def dur_to_str(hrs):
    if hrs is None: return '-'
    h=int(hrs); mn=int(round((hrs-h)*60))
    return f'{h}:{mn:02d}'

print('📂 يقرأ الملف ...')
df=pd.read_excel(INPUT_FILE,dtype=str,engine='openpyxl').fillna('-')
df=df.sort_values(['الرقم الوظيفي','التاريخ']).reset_index(drop=True)

# ── 1. تنظيف الساعات المطلوبة (مع استبدال 1 و2 بساعات النوبة) ─
df['الساعات المطلوبه']=df.apply(
    lambda r: clean_required_hours(r['الساعات المطلوبه'], r['وقت النوبة']), axis=1
)

# ── 2. حذف عمودي وقت الدخول.1 و وقت الخروج.1 ──────────────
for col in ['وقت الدخول.1','وقت الخروج.1']:
    if col in df.columns:
        df.drop(columns=[col],inplace=True)

# ── 3. تعبئة الأوقات الناقصة لكل حاضر وغير مكتمل ────────────
for idx,row in df.iterrows():
    status=row['الحالة']
    if status not in ('حاضر','غير مكتمل'):
        continue

    # غير الحالة إلى حاضر
    if status=='غير مكتمل':
        df.at[idx,'الحالة']='حاضر'

    entry=row['وقت الدخول']
    exit_=row['وقت الخروج']
    has_entry=parse_time(entry) is not None
    has_exit =parse_time(exit_) is not None

    if has_entry and has_exit:
        continue

    emp_id=row['الرقم الوظيفي']

    # أقرب يوم حضور سابق بأوقات كاملة
    mask=(df['الرقم الوظيفي']==emp_id) & \
         (df['الحالة'].isin(['حاضر','غير مكتمل'])) & \
         (df['وقت الدخول']!='-') & (df['وقت الخروج']!='-')

    prev=df[mask & (df['التاريخ']<row['التاريخ'])]
    ref=prev.iloc[-1] if len(prev)>0 else None
    if ref is None:
        nxt=df[mask & (df['التاريخ']>row['التاريخ'])]
        ref=nxt.iloc[0] if len(nxt)>0 else None

    # وقت النوبة كمصدر
    shift_str=str(row['وقت النوبة']).strip().replace('صباحًا','AM').replace('مساءً','PM').replace('صباحا','AM').replace('مساءا','PM')
    sp=re.split(r'\s*-\s*',shift_str)
    sh_start=sp[0].strip() if len(sp)==2 and parse_time(sp[0]) else None
    sh_end  =sp[1].strip() if len(sp)==2 and parse_time(sp[1]) else None

    if not has_entry and not has_exit:
        if sh_start:
            df.at[idx,'وقت الدخول']=to_arabic_time(sh_start)
            df.at[idx,'وقت الخروج']=to_arabic_time(sh_end) if sh_end else '-'
        elif ref is not None:
            df.at[idx,'وقت الدخول']=ref['وقت الدخول']
            df.at[idx,'وقت الخروج']=ref['وقت الخروج']
    elif not has_exit:
        if sh_end:
            df.at[idx,'وقت الخروج']=to_arabic_time(sh_end)
        elif ref is not None:
            df.at[idx,'وقت الخروج']=ref['وقت الخروج']
    elif not has_entry:
        if sh_start:
            df.at[idx,'وقت الدخول']=to_arabic_time(sh_start)
        elif ref is not None:
            df.at[idx,'وقت الدخول']=ref['وقت الدخول']


# ── 4b. ملء وقت النوبة والساعات المطلوبه لأي صف عنده أوقات لكن النوبة فاضية ──
for idx,row in df.iterrows():
    if row['وقت النوبة'] != '-': continue
    if parse_time(row['وقت الدخول']) is None: continue  # ما عنده أوقات أصلاً

    emp_id = row['الرقم الوظيفي']
    # أقرب يوم لنفس الموظف عنده وقت نوبة
    mask = (df['الرقم الوظيفي']==emp_id) & (df['وقت النوبة']!='-')
    prev = df[mask & (df['التاريخ']<row['التاريخ'])]
    ref  = prev.iloc[-1] if len(prev)>0 else None
    if ref is None:
        nxt = df[mask & (df['التاريخ']>row['التاريخ'])]
        ref = nxt.iloc[0] if len(nxt)>0 else None
    if ref is not None:
        df.at[idx,'وقت النوبة']       = ref['وقت النوبة']
        df.at[idx,'الساعات المطلوبه'] = ref['الساعات المطلوبه']

# ── 4. ملء مكان النوبة الفاضي لجميع الصفوف ──────────────────
for emp_id, grp in df.groupby('الرقم الوظيفي'):
    places=grp[grp['مكان النوبة']!='-']['مكان النوبة']
    if len(places)==0: continue
    most_common=places.mode()[0]
    df.loc[(df['الرقم الوظيفي']==emp_id) & (df['مكان النوبة']=='-'), 'مكان النوبة']=most_common

# ── 5. تصحيح إجمالي ساعات العمل ────────────────────────────
for idx,row in df.iterrows():
    ah=actual_hrs(row['وقت الدخول'],row['وقت الخروج'])
    if ah is not None:
        df.at[idx,'إجمالي ساعات العمل']=dur_to_str(ah)

# ── 6. ساعات النوبة لكل صف ───────────────────────────────────
df['_shift_hrs']=df['وقت النوبة'].apply(parse_shift_hours)
emp_avg=(df[df['_shift_hrs'].notna()].groupby('الرقم الوظيفي')['_shift_hrs'].mean().round(2))

def get_shift_hrs(row):
    if row['الحالة'] in LEAVE_STATUSES|REST_STATUSES:
        ah=actual_hrs(row['وقت الدخول'],row['وقت الخروج'])
        if ah is not None: return ah
    if row['_shift_hrs'] is not None: return row['_shift_hrs']
    return emp_avg.get(row['الرقم الوظيفي'],0)

df['_shift_hrs_filled']=df.apply(get_shift_hrs,axis=1)
df['_type']=df['الحالة'].apply(classify)

df['تأخير الدخول (د)']=df['الوصول متأخر'].apply(parse_duration_mins).apply(lambda x:x if x>=16 else 0)
df['مغادرة مبكرة (د)']=df['المغادرة مبكرا'].apply(parse_duration_mins).apply(lambda x:x if x>=16 else 0)

# ── 7. ملخص لكل موظف ─────────────────────────────────────────
def emp_summary(grp):
    hrs,types=grp['_shift_hrs_filled'],grp['_type']
    hp=hrs[types=='present'].sum()
    ha=hrs[types=='absent'].sum()
    hr=hrs[types=='rest'].sum()
    hl=hrs[types=='leave'].sum()
    hact=hp+hr+hl
    tot=hact+ha
    return pd.Series({
        'ساعات الدوام الفعلية':round(hact,2),
        'ساعات الغياب':        round(ha,2),
        'أيام الحضور':         int((types=='present').sum()),
        'أيام الغياب':         int((types=='absent').sum()),
        'أيام الراحة':         int((types=='rest').sum()),
        'أيام الإجازة':        int((types=='leave').sum()),
        'نسبة الحضور %':       round(hact/tot*100,1) if tot>0 else 0,
        'نسبة الغياب %':       round(ha/tot*100,1)   if tot>0 else 0,
    })

summary=df.groupby('الرقم الوظيفي').apply(emp_summary).reset_index()
df=df.merge(summary,on='الرقم الوظيفي',how='left')

# ── 8. ترتيب الأعمدة ─────────────────────────────────────────
orig_cols=[
    'الرقم الوظيفي','اسم الموظف','التاريخ','الساعات المطلوبه',
    'وقت النوبة','وقت الدخول','وقت الخروج','مكان النوبة','الحالة',
    'إجمالي ساعات العمل','إجمالي ساعات العمل داخل الدوام',
    'الوصول متأخر','المغادرة مبكرا','المغادرة مبكرا بعذر',
    'ساعات العمل الاضافيه','ساعات العمل الاضافية المؤكدة',
]
new_cols=[
    'ساعات الدوام الفعلية','ساعات الغياب',
    'تأخير الدخول (د)','مغادرة مبكرة (د)',
    'أيام الحضور','أيام الغياب','أيام الراحة','أيام الإجازة',
    'نسبة الحضور %','نسبة الغياب %'
]
final_cols=[c for c in orig_cols if c in df.columns]+[c for c in new_cols if c in df.columns]
df[final_cols].to_excel(OUTPUT_FILE,index=False,sheet_name='Attendance Details Report')

# ── 9. تنسيق ─────────────────────────────────────────────────
print('🎨 يضيف التنسيق ...')
wb=load_workbook(OUTPUT_FILE)
ws=wb.active

col_cat={
    'ساعات الدوام الفعلية':'blue','ساعات الغياب':'blue',
    'تأخير الدخول (د)':'blue','مغادرة مبكرة (د)':'blue',
    'أيام الحضور':'green','أيام الغياب':'green',
    'أيام الراحة':'green','أيام الإجازة':'green',
    'نسبة الحضور %':'green','نسبة الغياب %':'green',
}
fills={
    'orig': PatternFill('solid',start_color='1F4E79',end_color='1F4E79'),
    'blue': PatternFill('solid',start_color='2E75B6',end_color='2E75B6'),
    'green':PatternFill('solid',start_color='375623',end_color='375623'),
}
row_fills={
    'absent':PatternFill('solid',start_color='FFE0E0',end_color='FFE0E0'),
    'leave': PatternFill('solid',start_color='FFF3CD',end_color='FFF3CD'),
    'rest':  PatternFill('solid',start_color='E8F4FD',end_color='E8F4FD'),
}
hdr_font =Font(bold=True,color='FFFFFF',name='Arial',size=10)
data_font=Font(name='Arial',size=9)
center   =Alignment(horizontal='center',vertical='center',wrap_text=True)

status_col_idx=None
for i,cell in enumerate(ws[1],1):
    v=str(cell.value or '')
    cell.fill     =fills.get(col_cat.get(v,'orig'),fills['orig'])
    cell.font     =hdr_font
    cell.alignment=center
    ws.column_dimensions[get_column_letter(i)].width=18
    if v=='الحالة': status_col_idx=i

for row in ws.iter_rows(min_row=2):
    status=str(row[status_col_idx-1].value or '').strip() if status_col_idx else ''
    if status in ABSENT_STATUSES:  rf=row_fills['absent']
    elif status in LEAVE_STATUSES: rf=row_fills['leave']
    elif status in REST_STATUSES:  rf=row_fills['rest']
    else:                          rf=None
    for cell in row:
        cell.font     =data_font
        cell.alignment=center
        if rf: cell.fill=rf

ws.freeze_panes          ='A2'
ws.sheet_view.rightToLeft=True
wb.save(OUTPUT_FILE)

print(f'✅ تم! افتحي ملف: {OUTPUT_FILE}')
print(f'👥 موظفين: {df["الرقم الوظيفي"].nunique()} | صفوف: {len(df)}')
